import requests
from bs4 import BeautifulSoup
import openpyxl
import nltk
from nltk.tokenize import word_tokenize
from nltk import ne_chunk, pos_tag

class WebSleuth:
    def __init__(self):
        self.url = None
        self.emails = []
        self.urls = []
        self.phone_numbers = []
        self.names = []
        self.urls = []
        self.documents = []
        self.other = []
        self.harvested = []

    def nltk_test(self):
        try:
            nltk.data.find('tokenizers/punkt')
            nltk.data.find('chunkers/maxent_ne_chunker')
            nltk.data.find('taggers/averaged_perceptron_tagger')
            nltk.data.find('corpora/words')
            nltk.data.find('tokenizers/punkt_tab')
            nltk.data.find('taggers/averaged_perceptron_tagger_eng')
            nltk.data.find('chunkers/maxent_ne_chunker_tab')
        except LookupError:
            nltk.download('punkt')
            nltk.download('averaged_perceptron_tagger')
            nltk.download('maxent_ne_chunker')
            nltk.download('words')
            nltk.download('punkt_tab')
            nltk.download('averaged_perceptron_tagger_eng')
            nltk.download('maxent_ne_chunker_tab')
        print("NLTK packages are downloaded and ready to use.")

    def get_url(self):
        self.url = input("Enter a URL: ")

    def retrieve_website(self):
        print(f"Retrieving the website {self.url}")
        response = requests.get(self.url)
        if response.status_code == 200:
            print("Website retrieved successfully.")
            soup = BeautifulSoup(response.content, 'html.parser')
            nohtml = soup.get_text()
            urls = soup.find_all('a')
            #print("URLs found in the page:")
            #tokenize the text
            words = word_tokenize(nohtml)
            #tag the words
            tagged = pos_tag(words)
            #print(tagged)
            #named entities
            named_entities = ne_chunk(tagged)
            #print(named_entities)
            for name in named_entities:
                if hasattr(name, 'label'):#if the name has a label
                    self.names.append(name)

            for u in urls:
                # Sorting out the URLs
                if u.get('href') == None or u.get('href' in self.harvested):  # If the URL is empty
                    pass                    
                elif 'https://www.' in u.get('href'):  # If the URL is a full URL
                    self.urls.append(u.get('href'))
                elif '@' in u.get('href'):  # If the URL is an email
                    self.emails.append(u.get('href'))
                elif 'tel:' in u.get('href'):  # If the URL is a phone number
                    self.phone_numbers.append(u.get('href'))
                elif '.pdf' in u.get('href'):
                    self.documents.append(u.get('href'))
                else:  # If the URL is a partial URL
                    self.other.append(u.get('href')) 
                self.harvested.append(u.get('href'))
            
        else:
            print("Failed to retrieve the website")


    def SecondScan(self):
        for u in self.urls:
            if self.url not in u:
                pass
            else:
                print(f"Retrieving the website {u}")
                response = requests.get(u)
                if response.status_code == 200:
                    print("Website retrieved successfully.")
                    soup = BeautifulSoup(response.content, 'html.parser')
                    nohtml = soup.get_text()
                    urls = soup.find_all('a')
                    print("URLs found")
                    #tokenize the text
                    words = word_tokenize(nohtml)
                    print('Words tokenized')
                    #tag the words
                    tagged = pos_tag(words)
                    print('Words tagged')
                    #print(tagged)
                    #named entities
                    named_entities = ne_chunk(tagged)
                    print('Named entities found')
                    #print(named_entities)
                    for name in named_entities:
                        if hasattr(name, 'label'):#if the name has a label
                            self.names.append(name)

                for u in urls:
                    # Sorting out the URLs
                    if u.get('href') == None or u.get('href') in self.harvested: # If the URL is empty
                        pass
                    elif 'https://www.' in u.get('href'): # If the URL is a full URL
                        self.urls.append(u.get('href'))
                    elif '@' in u.get('href'): # If the URL is an email
                        self.emails.append(u.get('href'))
                    elif 'tel:' in u.get('href'): # If the URL is a phone number
                        self.phone_numbers.append(u.get('href'))
                    elif '.pdf' in u.get('href'):
                        self.documents.append(u.get('href'))
                    else: # If the URL is a partial URL
                        self.other.append(u.get('href'))
                    self.harvested.append(u.get('href'))
    
    def file_writer(self):
        # Create a new workbook
        workbook = openpyxl.Workbook()

        # Create worksheets for URLs, phone numbers, emails, and files
        url_sheet = workbook.create_sheet(title="URLs")
        phone_sheet = workbook.create_sheet(title="Phone Numbers")
        email_sheet = workbook.create_sheet(title="Emails")
        file_sheet = workbook.create_sheet(title="Files")
        name_sheet = workbook.create_sheet(title="Names")

        # Write data to the worksheets
        for url in self.urls:
            url_sheet.append([url])

        for phone_number in self.phone_numbers:
            phone_sheet.append([phone_number])

        for email in self.emails:
            email_sheet.append([email])

        for file in self.documents:
            file_sheet.append([file])
        
        for name in self.names:
            name_str = str(name)
            name_sheet.append([name_str])

        # Save the workbook
        workbook.save("output.xlsx")

    def main(self):
        self.nltk_test()
        self.get_url()
        self.retrieve_website()
        self.SecondScan()
        self.file_writer()


if __name__ == "__main__":
    sleuth = WebSleuth()
    sleuth.main()